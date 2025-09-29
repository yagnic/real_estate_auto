import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from datetime import datetime
import os

class DealAppraisalExcelGenerator:
    def __init__(self, deal_data):
        """Initialize with deal data dictionary"""
        self.deal_data = deal_data
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "Deal Appraisal"
        self.current_row = 1
        
        # Define styles
        self._setup_styles()
    
    def _setup_styles(self):
        """Setup Excel styles"""
        # Header style
        self.header_style = NamedStyle(name="header")
        self.header_style.font = Font(bold=True, color="FFFFFF", size=12)
        self.header_style.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_style.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        self.header_style.alignment = Alignment(horizontal="center", vertical="center")
        
        # Section title style
        self.section_style = NamedStyle(name="section")
        self.section_style.font = Font(bold=True, size=14, color="000080")
        self.section_style.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        
        # Data style
        self.data_style = NamedStyle(name="data")
        self.data_style.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Currency style
        self.currency_style = NamedStyle(name="currency")
        self.currency_style.number_format = '£#,##0'
        self.currency_style.border = self.data_style.border
        
        # Percentage style
        self.percent_style = NamedStyle(name="percent")
        self.percent_style.number_format = '0.0%'
        self.percent_style.border = self.data_style.border
        
        # Add styles to workbook
        if "header" not in self.wb.named_styles:
            self.wb.add_named_style(self.header_style)
        if "section" not in self.wb.named_styles:
            self.wb.add_named_style(self.section_style)
        if "data" not in self.wb.named_styles:
            self.wb.add_named_style(self.data_style)
        if "currency" not in self.wb.named_styles:
            self.wb.add_named_style(self.currency_style)
        if "percent" not in self.wb.named_styles:
            self.wb.add_named_style(self.percent_style)
    
    def _safe_get(self, value, default=None):
        """Safely get value, return default if None"""
        return value if value is not None else default
    
    def _safe_percent(self, value, decimals=1):
        """Safely convert decimal to percentage string"""
        if value is None:
            return "N/A"
        try:
            return f"{float(value)*100:.{decimals}f}%"
        except (TypeError, ValueError):
            return "N/A"
    
    def _safe_currency(self, value):
        """Safely format currency"""
        if value is None:
            return "N/A"
        try:
            return f"£{float(value):,.0f}"
        except (TypeError, ValueError):
            return "N/A"
    
    def _safe_number(self, value, decimals=0):
        """Safely format number"""
        if value is None:
            return "N/A"
        try:
            return f"{float(value):,.{decimals}f}"
        except (TypeError, ValueError):
            return "N/A"
    
    def _write_cell(self, row, col, value, style=None, merge_range=None):
        """Write a cell with optional styling and merging"""
        cell = self.ws.cell(row=row, column=col, value=value)
        if style:
            cell.style = style
        if merge_range:
            self.ws.merge_cells(f"{get_column_letter(col)}{row}:{get_column_letter(merge_range)}{row}")
        return cell
    
    def _format_value(self, value, value_type="auto"):
        """Format value based on type"""
        if value is None:
            return "N/A"
        
        if value_type == "currency":
            return self._safe_currency(value)
        elif value_type == "percent":
            return self._safe_percent(value)
        elif value_type == "number":
            return self._safe_number(value)
        else:
            return str(value) if value is not None else "N/A"
    
    def _add_section_header(self, title):
        """Add a section header"""
        self._write_cell(self.current_row, 1, title, "section", merge_range=8)
        self.current_row += 1
        return self.current_row - 1
    
    def _add_data_row(self, label, value, value_type="auto"):
        """Add a data row with label and value"""
        self._write_cell(self.current_row, 1, label, "data")
        
        if value_type == "currency" and isinstance(value, (int, float)) and value is not None:
            cell = self._write_cell(self.current_row, 2, value, "currency")
        elif value_type == "percent" and isinstance(value, (int, float)) and value is not None:
            cell = self._write_cell(self.current_row, 2, value, "percent")
        else:
            cell = self._write_cell(self.current_row, 2, self._format_value(value, value_type), "data")
        
        self.current_row += 1
        return cell
    
    def generate_excel(self, filename=None):
        """Generate the complete Excel file"""
        # Title
        self._write_cell(1, 1, "REAL ESTATE DEAL APPRAISAL", merge_range=8)
        self.ws.cell(1, 1).font = Font(bold=True, size=16)
        self.current_row = 3
        
        # Generate timestamp
        self._write_cell(self.current_row, 1, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        self.current_row += 2
        
        # Generate sections
        self._generate_deal_basics()
        self._generate_gdv_section()
        self._generate_costs_summary()
        self._generate_finance_analysis()
        self._generate_kpis()
        self._generate_funding_workings()
        
        # Auto-adjust column widths
        self._adjust_column_widths()
        
        # Save file
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"deal_appraisal_{timestamp}.xlsx"
        
        self.wb.save(filename)
        return filename
    
    def _generate_deal_basics(self):
        """Generate deal basics section"""
        self._add_section_header("DEAL BASICS")
        
        # KPI data for basic info
        kpis = self.deal_data.get('kpis', {})
        
        self._add_data_row("Deal Type", kpis.get('deal_type', 'Unknown'))
        
        timeline = kpis.get('timeline_months')
        self._add_data_row("Timeline", f"{timeline} months" if timeline else "N/A")
        
        self._add_data_row("Higher Risk Building", "Yes" if kpis.get('higher_risk_building') else "No")
        
        distance = kpis.get('travel_distance_miles')
        self._add_data_row("Travel Distance", f"{distance} miles" if distance else "N/A")
        
        self._add_data_row("Car Travel Time", kpis.get('car_travel_time', 'N/A'))
        
        self.current_row += 1
    
    def _generate_gdv_section(self):
        """Generate GDV section"""
        self._add_section_header("GROSS DEVELOPMENT VALUE (GDV)")
        
        gdv_data = self.deal_data.get('gdv', [])
        
        if gdv_data:
            # Headers
            headers = ["Market Type", "Build Type", "Floor", "Accommodation Type", 
                      "No of Units", "Avg Sqft/Unit", "Price/Unit", "Amount"]
            
            for col, header in enumerate(headers, 1):
                self._write_cell(self.current_row, col, header, "header")
            
            self.current_row += 1
            
            # Data rows (exclude the total row for now)
            unit_rows = [item for item in gdv_data if item.get('market_type') != 'Residential - Total']
            total_row = next((item for item in gdv_data if item.get('market_type') == 'Residential - Total'), None)
            
            for item in unit_rows:
                self._write_cell(self.current_row, 1, item.get('market_type', ''), "data")
                self._write_cell(self.current_row, 2, item.get('build_type', ''), "data")
                self._write_cell(self.current_row, 3, item.get('floor', ''), "data")
                self._write_cell(self.current_row, 4, item.get('accommodation_type', ''), "data")
                self._write_cell(self.current_row, 5, self._safe_get(item.get('no_of_units'), 0), "data")
                self._write_cell(self.current_row, 6, self._safe_get(item.get('average_sqft_per_unit'), 0), "data")
                
                price_per_unit = item.get('price_per_unit')
                if price_per_unit is not None:
                    self._write_cell(self.current_row, 7, price_per_unit, "currency")
                else:
                    self._write_cell(self.current_row, 7, "TBD", "data")
                
                amount = item.get('amount')
                if amount is not None:
                    self._write_cell(self.current_row, 8, amount, "currency")
                else:
                    self._write_cell(self.current_row, 8, "TBD", "data")
                
                self.current_row += 1
            
            # Total row
            if total_row:
                self._write_cell(self.current_row, 1, "TOTAL", "header")
                self._write_cell(self.current_row, 5, self._safe_get(total_row.get('no_of_units'), 0), "header")
                self._write_cell(self.current_row, 6, self._safe_get(total_row.get('average_sqft_per_unit'), 0), "header")
                
                total_amount = total_row.get('amount')
                if total_amount is not None:
                    self._write_cell(self.current_row, 8, total_amount, "currency")
                else:
                    self._write_cell(self.current_row, 8, "TBD", "header")
        
        else:
            self._add_data_row("GDV Data", "No data available")
        
        self.current_row += 2
    
    def _generate_costs_summary(self):
        """Generate costs summary section"""
        self._add_section_header("COSTS SUMMARY")
        
        # Acquisition costs
        acquisition = self.deal_data.get('acquisition_costs', {})
        asking_price = acquisition.get('asking_price')
        if asking_price is not None:
            self._add_data_row("Asking Price", asking_price, "currency")
        else:
            self._add_data_row("Asking Price", "TBD")
        
        # Build costs
        build_costs = self.deal_data.get('build_costs', {})
        self._add_data_row("Build Costs Total", build_costs.get('build_costs_total'), "currency")
        self._add_data_row("Cost per m²", build_costs.get('price_per_m2'), "currency")
        
        # Handle None for build_contingency_percent
        build_contingency = build_costs.get('build_contingency_percent')
        self._add_data_row("Build Contingency", self._safe_percent(build_contingency))
        
        # Statutory costs
        statutory = self.deal_data.get('statutory_costs', {})
        self._add_data_row("Statutory Costs Total", statutory.get('statutory_costs_total'), "currency")
        self._add_data_row("Nutrient Neutrality", statutory.get('nutrient_neutrality'), "currency")
        self._add_data_row("CIL", statutory.get('cil'), "currency")
        
        # Total development costs
        total_dev = self.deal_data.get('total_development_costs', {})
        self._add_data_row("Total Development Costs", total_dev.get('total_development_costs'), "currency")
        
        self.current_row += 1
    
    def _generate_finance_analysis(self):
        """Generate finance analysis section"""
        self._add_section_header("FINANCE ANALYSIS")
        
        # Finance costs
        finance = self.deal_data.get('finance_costs', {})
        
        self._add_data_row("Land LTV", self._safe_percent(finance.get('land_ltv'), decimals=0))
        self._add_data_row("Development LTV", self._safe_percent(finance.get('development_ltv'), decimals=0))
        self._add_data_row("Land Interest Rate", self._safe_percent(finance.get('land_interest_annual'), decimals=2))
        
        # Lenders costs
        lenders = self.deal_data.get('lenders_other_costs', {})
        self._add_data_row("Lenders' Other Costs", lenders.get('lenders_other_costs_total'), "currency")
        
        # Total funding costs
        funding = self.deal_data.get('total_funding_costs', {})
        self._add_data_row("Total Funding Costs", funding.get('funding_costs_total'), "currency")
        
        self.current_row += 1
    
    def _generate_kpis(self):
        """Generate KPIs section"""
        self._add_section_header("KEY PERFORMANCE INDICATORS")
        
        kpis = self.deal_data.get('kpis', {})
        net_profit = self.deal_data.get('net_profit_analysis', {})
        
        self._add_data_row("Net Profit", net_profit.get('net_profit'), "currency")
        self._add_data_row("Net Profit on GDV", net_profit.get('net_profit_on_gdv'), "percent")
        self._add_data_row("Own Funds Needed", kpis.get('own_funds_needed'), "currency")
        
        # Handle None values for percentages
        roi = kpis.get('return_on_own_funds')
        self._add_data_row("Return on Own Funds", roi if roi is not None else None, "percent")
        
        roi_pa = kpis.get('return_on_own_funds_per_annum')
        self._add_data_row("Return on Own Funds (per annum)", roi_pa if roi_pa is not None else None, "percent")
        
        self._add_data_row("LTC Criteria", kpis.get('lending_criteria_ltc', 'Unknown'))
        self._add_data_row("LTGDV Criteria", kpis.get('lending_criteria_ltgdv', 'Unknown'))
        
        # Post development
        post_dev = self.deal_data.get('post_development', {})
        self._add_data_row("Yearly Net Cashflow", post_dev.get('net_cashflow_per_annum'), "currency")
        
        annual_yield = kpis.get('annual_yield_on_own_funds')
        self._add_data_row("Annual Yield on Own Funds", annual_yield if annual_yield is not None else None, "percent")
        
        self.current_row += 1
    
    def _generate_funding_workings(self):
        """Generate funding workings section"""
        self._add_section_header("FUNDING WORKINGS")
        
        # Headers
        headers = ["Item", "Own Funds", "Borrowed", "Total"]
        for col, header in enumerate(headers, 1):
            self._write_cell(self.current_row, col, header, "header")
        self.current_row += 1
        
        # Data
        funding_workings = self.deal_data.get('funding_workings', {})
        
        items = [
            ("Asking Price", "asking_price"),
            ("Stamp Duty", "stamp_duty"),
            ("Sourcing Fee", "sourcing_fee"),
            ("Building Insurance", "building_insurance"),
            ("Legal Costs", "legal_professional_costs"),
            ("Acquisition Costs Total", "acquisition_costs_total"),
            ("Build Costs Total", "build_costs_total"),
            ("Professional Fees Total", "professional_fees_total"),
            ("Statutory Costs Total", "statutory_costs_total"),
            ("Funding Costs Total", "funding_costs_total"),
            ("Selling Costs Total", "selling_costs_total"),
            ("TOTAL COSTS", "total_costs")
        ]
        
        for label, key in items:
            item_data = funding_workings.get(key, {})
            
            self._write_cell(self.current_row, 1, label, "data")
            
            # Safely get values with default 0
            own = self._safe_get(item_data.get('own'), 0)
            borrow = self._safe_get(item_data.get('borrow'), 0)
            total = self._safe_get(item_data.get('total'), 0)
            
            self._write_cell(self.current_row, 2, own, "currency")
            self._write_cell(self.current_row, 3, borrow, "currency")
            self._write_cell(self.current_row, 4, total, "currency")
            
            # Highlight total row
            if label == "TOTAL COSTS":
                for col in range(1, 5):
                    self.ws.cell(self.current_row, col).style = "header"
            
            self.current_row += 1
        
        self.current_row += 1
    
    def _adjust_column_widths(self):
        """Auto-adjust column widths"""
        for column in self.ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            self.ws.column_dimensions[column_letter].width = adjusted_width


def generate_deal_excel(deal_data, filename=None):
    """
    Main function to generate Excel file from deal data
    
    Args:
        deal_data: Dictionary containing all deal analysis data
        filename: Optional filename, will auto-generate if not provided
    
    Returns:
        str: Path to generated Excel file
    """
    generator = DealAppraisalExcelGenerator(deal_data)
    return generator.generate_excel(filename)